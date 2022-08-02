import datetime
import json
import re
import sys

import psycopg2 as psycopg2
import unidecode as unidecode
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
from datetime import *
from openpyxl import load_workbook, Workbook
import time
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)


def carrega_selenium():
    Sem_tela = Options()

    Sem_tela.headless = True

    driver = webdriver.Chrome(options=Sem_tela, executable_path='chromedriver.exe')

    # driver = webdriver.Chrome('chromedriver.exe')

    return driver


def chr_remove(old, to_remove):
    new_string = old
    for x in to_remove:
        new_string = new_string.replace(x, '')
    return new_string


def get_agentes():
    # SALVA AGENTES

    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/agents?per_page=100"

    headers = {'Accept': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    response = requests.get(base_url, auth=auth, headers=headers)

    agentes = response.json()

    with open('agentes.json', 'w', encoding='utf-8') as f:
        json.dump(agentes, f, ensure_ascii=False, indent=4)

    return agentes


def traducoes(chamado):
    if chamado['status'] == 2:
        chamado['status'] = 'Aberto'
    elif chamado['status'] == 3:
        chamado['status'] = 'Pendente'
    elif chamado['status'] == 4:
        chamado['status'] = 'Resolvido'
    else:
        chamado['status'] = 'Fechado'

    if chamado['priority'] == 1:
        chamado['priority'] = 'Baixa'
    elif chamado['priority'] == 2:
        chamado['priority'] = 'Média'
    elif chamado['priority'] == 3:
        chamado['priority'] = 'Alta'
    else:
        chamado['priority'] = 'Urgente'


def get_chamados():
    filter_ticket = 'https://fotosensores-dnit.freshdesk.com/api/v2/search/tickets'

    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

    headers = {'Accept': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    cols = ['ID do ticket', 'Status', 'Assunto', 'Prioridade', 'Tipo', 'Agente', 'Data de Criação', 'Tags', 'FOTOBOT']

    data = pd.DataFrame(columns=cols)

    print(f'Conectando ao Freshdesk')

    agentes = get_agentes()

    for i in range(1, 11):

        print(f'Salvando dados da pagina {i}')

        # response = requests.get(f'{base_url}?updated_since=2015-01-19&per_page=100&page={i}', auth=auth,
        #                         headers=headers)
        #
        # link = response.headers.get("link")
        #
        # retry = response.headers.get("Retry-After")

        # if retry is not None:
        #     int_retry = int(retry) + 1
        #     print(f'esperando {int_retry} segundos')
        #     time.sleep(int_retry)
        #     response = requests.get(f'{base_url}?updated_since=2015-01-19&per_page=100&page={i}', auth=auth,
        #                             headers=headers)
        #
        #     link = response.headers.get("link")
        #
        #     print(f'Tentando novamente na pagina {i}')

        response = requests.get(f'{filter_ticket}?query="status:2%20OR%20status:3"&page={i}', auth=auth,
                                headers=headers)

        response = response.json()

        # substitui responder_id por nome do contato

        for chamado in response['results']:
            tz = timedelta(hours=3)
            chamado['created_at'] = datetime.strptime(chamado['created_at'], "%Y-%m-%dT%H:%M:%Sz")
            chamado['created_at'] = chamado['created_at'] - tz
            chamado['Dias_aberto'] = datetime.today() - chamado['created_at']

            for agente in agentes:
                if agente['id'] == chamado['responder_id']:
                    chamado['responder_id'] = agente['contact']['name']

                if chamado['responder_id'] is None:
                    chamado['responder_id'] = 'Não Atribuido'

            traducoes(chamado)

            # if chamado['tags']:
            #     chamado['tags'] = chamado['tags'][0]
            # else:
            #     chamado['tags'] = None

            ultima_att = datetime.today()
            ultima_att = f'{ultima_att.day}/{ultima_att.month}/{ultima_att.year} {ultima_att.hour}:{ultima_att.minute}'
            data = data.append({
                'ID do ticket': chamado['id'],
                'Status': chamado['status'],
                'Prioridade': chamado['priority'],
                'Assunto': chamado['subject'],
                'Tipo': chamado['type'],
                'Agente': chamado['responder_id'],
                'Data de Criação': chamado['created_at'],
                'Dias aberto': chamado['Dias_aberto'],
                'Tags': chamado['tags'],
                'Peças': chamado['custom_fields']['cf_peas62992'],
                'Ultima_att': ultima_att,
                'FOTOBOT': chamado['custom_fields']['cf_fotobot']

            }, ignore_index=True)

    return data


def create_ticket(titulo, descricao, contato):
    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

    headers = {'Content-Type': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    exemple = {
        'subject': titulo,
        'description': descricao,
        'email': contato,
        'priority': 4,
        'status': 2,
        'group_id': 67000215926,
        'responder_id': 67033345228,
        'type': 'Equipamento Offline'
    }

    exemplo = json.dumps(exemple)

    response = requests.post(f'{base_url}', auth=auth,
                             headers=headers, data=exemplo)

    return response


def create_service_task(parent_id, titulo, descricao, local_servico, responsavel):
    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

    headers = {'Content-Type': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    hoje = date_iso_format()

    exemple = {
        "parent_id": parent_id,
        "type": "Service Task",
        "subject": titulo,
        "description": descricao,
        'requester_id': 67034342692,
        'group_id': 67000573741,
        'responder_id': responsavel,
        "status": 2,
        "priority": 4,
        "custom_fields": {
            "cf_fsm_contact_name": "FOTO BOT",
            "cf_fsm_phone_number": "54996794368",
            "cf_fsm_service_location": local_servico,
            "cf_fsm_appointment_start_time": hoje,
            "cf_fsm_appointment_end_time": hoje
        },
    }

    exemplo = json.dumps(exemple)

    response = requests.post(f'{base_url}', auth=auth,
                             headers=headers, data=exemplo)

    return response


def add_reply(ticket_id, anotacao):
    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

    headers = {'Content-Type': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    verificacao = datetime.today()

    data = f'{verificacao.day}/{verificacao.month}/{verificacao.year} {verificacao.hour}:{verificacao.minute} '

    anotacao = {

        'body': f'{anotacao} {data}'

    }

    response = requests.post(f'{base_url}/' + str(ticket_id) + '/reply', auth=auth,
                             headers=headers, data=json.dumps(anotacao))

    return response


def close_ticket(ticket_id):
    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

    headers = {'Content-Type': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    ticket = {

        'status': 4,
    }

    response = requests.put(f'{base_url}/' + str(ticket_id), auth=auth,
                            headers=headers, data=json.dumps(ticket))

    return response


def close_task(ticket_id):
    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

    headers = {'Content-Type': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    task = {
        'custom_fields': {

            'cf_peas62992': 'Sem necessidade'

        },

        'status': 4,

    }

    response = requests.put(f'{base_url}/' + str(ticket_id), auth=auth,
                            headers=headers, data=json.dumps(task))

    return response


def cria_lista_central():
    print('Carregando dados iniciais Central...')

    driver = carrega_selenium()

    # driver = webdriver.Chrome('chromedriver.exe')

    driver.get('http://fotosensores.velsis.com.br/CentralVelsis/login?39')

    # LOGIN NA PLATAFORMA

    print("Fazendo login em Velsis")

    login = driver.find_element(By.NAME, value='username')

    senha = driver.find_element(By.NAME, value='password')

    login.send_keys('fabio.silva@fotosensores.com')

    senha.send_keys('Gogeta96@')

    click = driver.find_element(By.NAME, value='entrar')

    click.click()

    # SELEÇÃO DE CONTRATO

    print("Selecionando Contrato!")

    click = driver.find_element(By.ID, value='btnSubmit')

    click.click()

    # ENTRAR MONITORAMENTO

    print("Entrando em Monitoramento!")

    click = driver.find_element(By.ID, value='menu.monitoramento')

    click.click()

    click = driver.find_element(By.ID, value='menuVAliveMonitoramento')

    click.click()

    # EXIBINDO CENTRAL DOS DOIS CONTRATOS

    print('Exibindo os dois contratos')

    driver.get('http://fotosensores.velsis.com.br/CentralVelsis/wicket/bookmarkable/br.com.velsis.central'
               '.monitoramento.valive.VAliveMonitoramento?749&tempoAtualizacao=15.0&filtroUltimosDias=0&contratos'
               '=%5B4,+1%5D&somenteAtivos=false')

    click = driver.find_element(By.XPATH, '/html/body/div/div[1]/div/div/div/form/div[3]/div')

    click.click()

    click = driver.find_element(By.XPATH, '/html/body/div/div[1]/div/div/div/form/div[3]/div/div/div/ul/li[1]')

    click.click()

    click = driver.find_element(By.XPATH, '/html/body/div/div[1]/div/div/div/form/div[3]/div/div/div/ul/li[2]')

    click.click()

    click = driver.find_element(By.XPATH, '/html/body/div/div[1]/div/div/div/form/button[2]/span')

    click.click()

    # MONITORAMENTO

    print("Recolhendo dados...")

    driver.get('http://fotosensores.velsis.com.br/CentralVelsis/wicket/bookmarkable/br.com.velsis.central'
               '.monitoramento.valive.VAliveMonitoramento?390&tempoAtualizacao=15.0&filtroUltimosDias=0&contratos'
               '=%5B1%5D&somenteAtivos=false')

    headers = []
    columns = dict()

    table_id = driver.find_element(By.ID, value='tblStatus')
    all_rows = table_id.find_elements(By.XPATH, "/html/body/div/div[4]/div/table/tbody/tr")

    # --- headers ---

    row = all_rows[0]
    all_items = row.find_elements(By.XPATH, "/html/body/div/div[4]/div/table/thead/tr[2]/th")
    for item in all_items:
        name = item.text
        columns[name] = []
        headers.append(name)

    print("Cabeçalho recebido...")

    print("Recebendo valores das linhas...")

    # --- data ---

    for row in all_rows:
        all_items = row.find_elements(By.TAG_NAME, value="td")
        for name, item in zip(headers, all_items):
            value = item.text
            columns[name].append(value)

    df = pd.DataFrame(columns)

    print("Todos os dados recebido!")

    return df


def excesoes_offlines():
    excel = load_workbook(r"C:\Users\User\Dropbox\Rede Fotosensores\Dashboard\excessoes.xlsx")
    planilha_excessoes = excel['excessoes']

    excessoes = {}

    for linha in planilha_excessoes.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True):
        serial = linha[0]

        excessoes[serial] = {
            'codigo': linha[1],
            'municipio': linha[2],
            'motivo': linha[3]
        }

    excel.close()

    return excessoes


def add_excessao_campo_fotobot():
    dno = []
    for ticket in tickets_abertos:
        if ticket[8] == "Não abrir outros tickets":
            dno.append(ticket[0])
        if ticket[8] == "Fechar este ticket e não abrir outros":
            dno.append(ticket[0])
            add_reply(ticket[2],
                      "Chamado será fechado e adicionado as excessões conforme orientado pelo campo FOTOBOT verificado em")
            close_task(ticket[2] + 1)
            close_ticket(ticket[2])

    excel = load_workbook(r"C:\Users\User\Dropbox\Rede Fotosensores\Dashboard\excessoes.xlsx")
    planilha_excessoes = excel['excessoes']
    excessoes = excesoes_offlines()
    for item in dno:
        if str(item) not in excessoes:
            excessao = [item]
            planilha_excessoes.append(excessao)
            excel.save(r"C:\Users\User\Dropbox\Rede Fotosensores\Dashboard\excessoes.xlsx")
            excel.close()
        else:
            print('Sem novas excessoes')


def verifica_tipo_ticket_se_e_relevante_para_deixar_equipamento_off(tipo):
    tipos_de_ticket = ['Equipamento Offline',
                       'Conectorização',
                       'Internet',
                       'Equipamento sem energia',
                       'Instalação / Reparo de energia eletrica',
                       'Manutenção Preventiva',
                       'Manutenção corretiva']

    tipo = str(tipo)
    r = ''
    for i in tipos_de_ticket:
        r = i in tipo
        if r == True:
            break

    return r


def verifica_assunto_e_relevante(assunto):
    assuntos_nao_relevantes = [

        'tachoes',
        'tachao',
        'iluminador'
        'sinalizacao',
        'horizontal',
        'vertical',
        'poda',
        'zebrado',
        'Tampa'
    ]

    assunto = unidecode.unidecode(assunto)

    r = True
    for i in assuntos_nao_relevantes:
        teste = assunto.find(i)
        if teste > -1:
            r = False
            break

    return r


def verifica_laco_rompido_no_assunto(subject):
    laco_rompido = [

        'rompido',
        'Laço',
        'Laco'
        'Rompido',
        'laco',
        'Laço rompido'

    ]

    assunto = unidecode.unidecode(subject)

    r = False
    for i in laco_rompido:
        teste = assunto.find(i)
        if teste > -1:
            r = True
            break

    return r


def atualiza_planilha():
    # CRIA AS PLANILHAS E SALVA EM ATUALIZACAO.XLSX

    offlines = cria_lista_central()
    chamados = get_chamados()

    with pd.ExcelWriter(r"C:\Users\User\Dropbox\Rede Fotosensores\Dashboard\Atualizacao.xlsx") as writer:
        chamados.to_excel(writer, sheet_name="tickets_freshdesk", index=False)
        offlines.to_excel(writer, sheet_name="offlines", index=False)


def date_iso_format():
    td = timedelta(hours=8)

    hoje = datetime.now() + td

    hoje = hoje.isoformat()

    return hoje


def get_responsavel_and_coordenadas():
    excel_banco = load_workbook('banco.xlsx')
    planilha = excel_banco['Google - Colar Valores']

    banco = {}

    agents = get_agentes()

    for i in planilha.iter_rows(min_row=2, min_col=0, max_col=15, values_only=True):

        id_tec_responsavel = 0

        for agente in agents:

            tecnico_freshdesk = unidecode.unidecode(agente['contact']['name'])
            tecnico_distribuicao_tecnica = unidecode.unidecode(i[14])

            if tecnico_freshdesk.find(tecnico_distribuicao_tecnica) > -1:
                id_tec_responsavel = agente['id']
                break
            elif i[14] == "SECURITEL":
                id_tec_responsavel = 67037224654
                break
            elif i[14] == "Clayton":
                id_tec_responsavel = 67032076394
                break

        coordenadas = '{}'.format(i[11].rstrip() + ',' + i[12].lstrip())

        serial_banco = i[6]

        banco[serial_banco] = {

            'coordenadas': coordenadas,

            'id_responsavel': id_tec_responsavel

        }

    return banco


def verifica_se_central_ta_toda_offline():
    if not onlines:
        sys.exit('Central caiu!')


def cria_planilha_offlines(offlines):
    cols = ['Serial', 'Municipio', 'Ticket', 'Cod Equipamento', 'KM', 'Status equipamento']

    data = pd.DataFrame(columns=cols)

    for offline in offlines:
        if offline[6] == []:
            ticket = 'Excessão'
        else:
            ticket = offline[6]
        serial = int(offline[0])
        municipio = offline[1]
        cod_equipamento = offline[2]
        br_km = offline[3]
        status_equipamento = offline[4]

        data = data.append({
            'Serial': serial,
            'Municipio': municipio,
            'Ticket': ticket,
            'Cod Equipamento': cod_equipamento,
            'KM': br_km,
            'Status equipamento': status_equipamento,
        }, ignore_index=True)

    with pd.ExcelWriter(r"C:\Users\User\Dropbox\Rede Fotosensores\Dashboard\offlines.xlsx") as writer:
        data.to_excel(writer, sheet_name="tickets_freshdesk", index=False)


while True:

    atualiza_planilha()

    # INICIANDO LISTAS
    offlines = []
    onlines = []
    tickets_abertos = []

    # LENDO PLANILHA ATUALIZACAO.XLSX
    excel = load_workbook(r"C:\Users\User\Dropbox\Rede Fotosensores\Dashboard\Atualizacao.xlsx")
    planilha_central = excel['offlines']
    planilha_tickets = excel['tickets_freshdesk']

    # BUSCANDO DADOS DOS EQUIPAMENTOS
    banco = get_responsavel_and_coordenadas()

    # VERIFICA PLANILHA CENTRAL E ADICIONA OS OFFLINES E OS ONLINES EM SUAS LISTAS COM OS INDICES SERIAL, MUNICIPIO E STATUS
    for linha in planilha_central.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True):

        status_equipamento = linha[10]

        serial_equipamento = linha[0].lstrip('0')

        municipio = linha[3]

        cod_equipamento = linha[1]

        br_km = linha[2]

        status_mlr = linha[8].find('RescueMode')

        if status_mlr > -1:
            status_mlr = True
        else:
            status_mlr = False

        if status_equipamento == 'Acessar Equipamento':

            onlines.append([serial_equipamento, municipio, cod_equipamento, br_km, status_equipamento, status_mlr])

        else:

            hora_atual = datetime.today()
            date_time_obj = datetime.strptime(status_equipamento, '%d/%m/%Y %H:%M:%S')
            offline_a_3_horas = hora_atual - date_time_obj > timedelta(hours=3)

            offlines.append(
                [serial_equipamento, municipio, cod_equipamento, br_km, status_equipamento, offline_a_3_horas])

    # VERIFICA SE A CENTRAL TA TODA OFFLINE COM BASE NA LISTA DE ONLINES ESTAR VAZIA
    verifica_se_central_ta_toda_offline()

    # VERIFICA PLANILHA DE TICKETS E ADICIONANDO OS TICKETS A LISTA COM OS INDICES SERIAL, STATUS, ID, TIPO, TAGS, VERIFICAÇÃO SE TEM 'EQUIPAMENTO OFFLINE' NO ASSUNTO
    for linha in planilha_tickets.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True):
        serial_tickets = linha[2][0:5]
        assunto = linha[2].lower()
        status = linha[1]
        id = linha[0]
        tipo = linha[4]
        tags = linha[7]
        pecas = linha[10]
        busca_assunto_equipamento_offline = assunto.find('equipamento offline')
        fotobot = linha[8]
        agente = linha[5]
        prioridade = linha[3]
        dias_aberto = linha[9]
        data_criacao = linha[6]

        equipamento_offline = False

        if busca_assunto_equipamento_offline > -1:
            equipamento_offline = True

        tickets_abertos.append(
            [serial_tickets, status, id, tipo, tags, equipamento_offline, pecas, assunto, fotobot, agente, prioridade,
             dias_aberto, data_criacao])

    # VERIFICA OS TICKETS ABERTOS COM O CAMPO NÃO ABRIR OUTROS TICKETS PARA ADCIONAR AS EXCESSOES
    add_excessao_campo_fotobot()

    # VERIFICA A LISTA DE OFFLINES E COMPARA COM A DE TICKETS PARA ENCONTRAR O TICKET RESPECTIVO E FAZ O VINCULO COM O ID DO TICKET
    for offline in offlines:
        offline.append([])
        for ticket in tickets_abertos:
            if offline[0] == ticket[0] and verifica_tipo_ticket_se_e_relevante_para_deixar_equipamento_off(
                    ticket[3]) and verifica_assunto_e_relevante(ticket[7]) \
                    or (offline[0] == ticket[0] and verifica_laco_rompido_no_assunto(ticket[7])
                        or (offline[0] == ticket[0] and ticket[8] == "Não fechar este ticket")
            ):
                offline[6] = ticket[2]

    # CRIANDO PLANILHA DOS OFFLINES ANTES DOS FILTROS DE EXCESSOES E TEMPO OFFLINE
    cria_planilha_offlines(offlines)

    # VERIFICA A LISTA DE ONLINE E COMPARA COM A DE TICKET PARA ENCONTRAR O TICKET RESPECTIVO E FAZ O VINCULO COM O ID DO TICKET
    for online in onlines:
        online.append([])
        for ticket in tickets_abertos:

            if ticket[8] == "Não fechar este ticket":
                break

            if ticket[3] == 'Equipamento Offline':
                if online[0] == ticket[0]:
                    online[6].append(ticket[2])

    # ADICIONA TAREFAS AOS EQUIPAMENTOS
    for online in onlines:
        online.append([])
        for ticket in tickets_abertos:
            if ticket[8] == "Não fechar este ticket":
                break
            if ticket[3] == 'Service Task' and ticket[5]:
                if online[0] == ticket[0]:
                    online[7].append(ticket[2])

    # VERIFICA SE O TICKET DO EQUIPAMENTO ONLINE ESTA COM O CAMPO PEÇA NECESSARIA
    for online in onlines:
        online.append([])
        for ticket in tickets_abertos:
            if online[0] == ticket[0]:
                if ticket[6] == 'Peça necessaria':
                    online[8] = True
                    break
                else:
                    online[8] = False

    # FILTRA OS OFFLINES COM BASE NA LISTA DE EXCESSOES
    for offline in offlines:
        offline.append([])
        offline[7] = False
        excessoes = excesoes_offlines()

        hora_atual = datetime.today()
        date_time_obj = datetime.strptime(offline[4], '%d/%m/%Y %H:%M:%S')
        offline_a_6_horas = hora_atual - date_time_obj > timedelta(hours=6)

        for excessao in excessoes:
            if excessoes[excessao]['motivo'] == "Fotovoltaico" and offline_a_6_horas:
                offline[7] = False

            elif str(excessao) == offline[0]:
                offline[7] = True
                break

    # CRIAÇÃO DE TICKETS PARA OS OFFLINES COM BASE NAS VERIFICAÇÕES
    for offline in offlines:

        ticket = offline[6]
        serial = int(offline[0])
        municipio = offline[1]
        tempo_offline = offline[5]
        excessao = offline[7]
        cod_equipamento = offline[2]
        br_km = offline[3]
        status_equipamento = offline[4]

        if not ticket and tempo_offline and not excessao:
            titulo = f'{serial} - {cod_equipamento} - {br_km} - {municipio} - Equipamento offline'
            descricao = f'Offline desde {status_equipamento}'
            contato = 'fabiogjo@hotmail.com'

            r = create_ticket(titulo, descricao, contato)

            id_ticket_criado = int(r.headers.get('Location')[-5:])

            r = create_service_task(id_ticket_criado, titulo, descricao, banco[serial]['coordenadas'],
                                    banco[serial]['id_responsavel'])

            if r.status_code == 201:
                print(f"Ticket {id_ticket_criado} criado para o equipamento {serial} - {municipio}")
            else:
                print(f"Falha ao criar ticket para o equipamento {serial} - {banco[serial]['id_responsavel']}")
                response = json.loads(r.content)
                print(response)

    # FECHAMENTO DE TICKETS PARA OS EQUIPAMENTOS QUE JA ESTAO ONLINE
    for online in onlines:
        tickets = online[6]
        tasks = online[7]
        serial = online[0]
        municipio = online[1]
        cod_equipamento = online[2]
        br_km = online[3]
        status_equipamento = online[4]
        peca_necessaria = online[8]
        rescue_mode = online[5]

        if tickets and not rescue_mode and not peca_necessaria:
            for task in tasks:
                add_reply(str(task), "Equipamento encontra-se online conforme verificação em")
                close_task(str(task))
                print(f'Adicionado anotação em tarefa {task} e status resolvido.')

            for ticket in tickets:
                add_reply(str(ticket), "Equipamento encontra-se online conforme verificação em")
                close_ticket(str(ticket))
                print(f'Adicionado anotação em ticket {ticket} e status resolvido.')

    atualiza_planilha()

    conn = psycopg2.connect(database="dnit_bi",
                            host="localhost",
                            user="postgres",
                            password="123456",
                            port="5432")

    cursor = conn.cursor()

    cursor.execute("DELETE FROM dnit_bi_offline")

    cursor.execute("DELETE FROM dnit_bi_ticket_freshdesk")


    for ticket in tickets_abertos:

        tipo = ticket[3]

        serial = ticket[0]

        id_ticket = ticket[2]

        print(id_ticket)

        status_equipamento = 0

        agente = ticket[9]

        tags = ticket[4]

        pecas = ticket[6]

        assunto = ticket[7]

        status = ticket[1]

        data_criacao = ticket[12]

        dias_aberto = ticket[11]

        prioridade = ticket[10]

        if serial == "Setor":
            continue

        cursor.execute(f"SELECT id FROM dnit_bi_equipamento WHERE numero_de_serie = {int(serial)}")

        id = cursor.fetchone()[0]

        cursor.execute(
            "INSERT INTO dnit_bi_ticket_freshdesk (id_ticket, tipo, data_criacao, tags, dias_aberto, pecas, "
            "prioridade, assunto, status, agente, equipamento_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (id_ticket, tipo, data_criacao, tags, dias_aberto, pecas, prioridade, assunto, status, agente, id))

        if tipo == 'Equipamento Offline':
            for offline in offlines:
                if offline[5]:
                    if offline[0] == serial:
                        status_equipamento = offline[4]
                        break
                    else:
                        status_equipamento = "01/01/1900 00:00:00"

            cursor.execute(f"SELECT id FROM dnit_bi_equipamento WHERE numero_de_serie = {int(serial)}")

            id = cursor.fetchone()[0]

            agora = datetime.now()

            dataobj = datetime.strptime(status_equipamento, '%d/%m/%Y %H:%M:%S')

            dias_offline = agora - dataobj

            cursor.execute(
                "INSERT INTO dnit_bi_offline (offline_desde, equipamento_id, ticket_id, dias_offline) VALUES (%s, %s, %s, %s)",
                (status_equipamento, id, id_ticket, dias_offline.days))

    conn.commit()

    conn.close()

    print('Dados inseridos no Banco de dados')

    print(f'Finalizada vericação as {datetime.now()}')

    print('proxima verificação em 10 minutos')

    time.sleep(600)
